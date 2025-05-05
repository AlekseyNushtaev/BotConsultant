import asyncio
import csv
import datetime
import openpyxl
from aiogram.utils.keyboard import InlineKeyboardBuilder

from bot import bot
from spread import get_sheet

from aiogram import Router, types, F
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart, StateFilter, ChatMemberUpdatedFilter, KICKED, MEMBER
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State, default_state
from aiogram.types import Message, CallbackQuery, ChatMemberUpdated, FSInputFile, InlineKeyboardButton, InputMediaPhoto, \
    InputMediaVideo, InputMediaDocument
from collections import defaultdict
from typing import Dict, List
from config import ADMIN_IDS
from db.util import add_user_to_db, update_user_blocked, update_user_unblocked, add_question_to_db, get_all_questions, \
    delete_all_questions, get_all_users, get_all_users_unblock
from keyboard import create_kb, kb_button

router = Router()

media_groups: Dict[str, List[Message]] = defaultdict(list)
timers: Dict[str, asyncio.Task] = {}
builder = InlineKeyboardBuilder()

builder.row(InlineKeyboardButton(text="Записаться на консультацию ✅", callback_data="quest_1"))
builder.row(InlineKeyboardButton(text="Подписаться на телеграм канал", url="https://t.me/andreikuvshinov"))

main_keyboard_markup = builder.as_markup()

class FSMFillForm(StatesGroup):
    get_full_name = State()
    get_question = State()
    get_contact = State()
    send = State()
    text_add_button = State()
    check_text_1 = State()
    check_text_2 = State()
    text_add_button_text = State()
    text_add_button_url = State()
    photo_add_button = State()
    check_photo_1 = State()
    check_photo_2 = State()
    photo_add_button_text = State()
    photo_add_button_url = State()
    video_add_button = State()
    check_video_1 = State()
    check_video_2 = State()
    video_add_button_text = State()
    video_add_button_url = State()
    check_video_note_1 = State()
    send_id = State()
    send_to_one = State()


async def scheduler(time):
    while True:
        await asyncio.sleep(10)
        print(datetime.datetime.now())
        try:
            sheet = await get_sheet()
            sheet.clear()
            quests = get_all_questions()
            users = get_all_users()
            sheet.append_rows(quests)
            sheet.append_rows(users)
            empty_row = ['' for cell in range(sheet.col_count)]
            sheet.insert_row(empty_row, len(quests) + 1)
            print(datetime.datetime.now())
        except Exception as e:
            await bot.send_message(1012882762, str(e))
        await asyncio.sleep(time)


async def process_media_group(media_group_id: str):
    await asyncio.sleep(5)  # Даем время на получение всех частей группы
    messages = media_groups.pop(media_group_id, [])

    if not messages:
        return

    # Собираем медиа для альбома
    media = []
    for idx, msg in enumerate(messages):
        if msg.photo:
            file_id = msg.photo[-1].file_id
            item = InputMediaPhoto(media=file_id)
        elif msg.video:
            file_id = msg.video.file_id
            item = InputMediaVideo(media=file_id)
        elif msg.document:
            file_id = msg.document.file_id
            item = InputMediaDocument(media=file_id)
        else:
            continue

        # Добавляем подпись только к первому элементу
        if idx == 0 and msg.caption:
            item.caption = msg.caption
            # Исправление здесь: проверяем наличие parse_mode
            if hasattr(msg, 'parse_mode'):
                item.parse_mode = msg.parse_mode

        media.append(item)

    if media:
        users = get_all_users()
        cnt = 0
        for user in users[1:]:
            if not user[6]:
                try:
                    await bot.send_media_group(chat_id=int(user[1]), media=media)
                    await asyncio.sleep(0.2)
                    cnt += 1
                except Exception as e:
                    print(f"Ошибка отправки медиагруппы: {e}")
        await bot.send_message(1012882762, f'Переслано {cnt} юзерам')


@router.message(CommandStart(), StateFilter(default_state))
async def process_start_user(message: Message):
    add_user_to_db(
        message.from_user.id,
        message.from_user.username,
        message.from_user.first_name,
        message.from_user.last_name,
        datetime.datetime.now()
    )
    await message.answer_video_note('DQACAgIAAxkBAAMkZ7YTobVOP3LcI_-weFilb18kwXkAAiJtAALpnalJe1wSU2ZnRM82BA')
    await asyncio.sleep(0.3)
    await message.answer(
        text="""
Здравствуйте! 👋🏻

Вы попали в официальный бот Андрея Кувшинова – бизнес-юриста.

🔹 Здесь вы можете оставить заявку на бесплатную юридическую консультацию по вопросам корпоративного права, банкротства и защиты бизнеса.

📌 Как оставить заявку?
Просто нажмите на кнопку ниже и заполните короткую форму – мы свяжемся с вами в ближайшее время.

💼 Помогаем бизнесу решать сложные юридические вопросы просто и эффективно!
        """,
        parse_mode=ParseMode.HTML,
        reply_markup=main_keyboard_markup
    )
    await asyncio.sleep(1)
    await message.answer(text="""
Мы в ЮК «СТАРТ» не просто юристы — мы ваши антикризисные союзники. Подготовили для вас 3 практичных гайда, которые помогут пережить налоговую проверку и не утонуть в кризисе.

📍 Первый гайд — антикризисное руководство для логистики: как договориться с лизингом и банками, что делать с дебиторкой, как не слить имущество и не угодить под субсидиарку.

📍 Второй — про налоговые проверки 2025: что ждать, как действовать, что говорить (а главное — чего не говорить), чтобы не оказаться в плане выездной.

📍 Третий гайд — что делать, если вызывают в налоговую или МВД, как отвечать на вопросы, чтобы не подставить себя или компанию, какие ошибки совершают даже опытные руководители и как их избежать.

🎁 Эти гайды — не вода, а конкретика и опыт. Скачивайте, читайте и действуйте. Пока другие в панике — вы уже на шаг впереди. Ссылки ниже 👇
    """,
                         reply_markup=create_kb(1,
                                                faq_1='Гайд "Как сохранить логистический бизнес"',
                                                faq_2='Гайд по Налоговым проверкам 2025',
                                                faq_3='Гайд "Как вести себя на допросе"'
                                                ))


@router.callback_query(F.data == 'quest_1', StateFilter(default_state))
async def step_1(cb: CallbackQuery, state: FSMContext):
    add_user_to_db(
        cb.from_user.id,
        cb.from_user.username,
        cb.from_user.first_name,
        cb.from_user.last_name,
        datetime.datetime.now()
    )
    await cb.message.answer(text="""
Отлично, теперь напишите как к Вам обращаться?      
    """)
    await state.set_state(FSMFillForm.get_full_name)


@router.callback_query(F.data == 'new', StateFilter(default_state))
async def step_1_1(cb: CallbackQuery, state: FSMContext):
    add_user_to_db(
        cb.from_user.id,
        cb.from_user.username,
        cb.from_user.first_name,
        cb.from_user.last_name,
        datetime.datetime.now()
    )
    await cb.message.answer(text="""
Отлично, теперь напишите как к Вам обращаться?     
    """)
    await state.set_state(FSMFillForm.get_full_name)


@router.message(F.text, StateFilter(FSMFillForm.get_full_name))
async def step_2(message: types.Message, state: FSMContext):
    await state.update_data(full_name=message.text)
    await message.answer(text='Отлично, какой у вас вопрос?')
    await state.set_state(FSMFillForm.get_question)


@router.message(F.text, StateFilter(FSMFillForm.get_question))
async def step_3(message: types.Message, state: FSMContext):
    await state.update_data(question=message.text)
    await message.answer(text='Оставьте свои контактные данные (номер телефона или email)',
                         reply_markup=create_kb(1, telegram="Связаться в телеграм"))
    await state.set_state(FSMFillForm.get_contact)


@router.message(F.text, StateFilter(FSMFillForm.get_contact))
async def step_4_1(message: types.Message, state: FSMContext):
    contact = message.text
    dct = await state.get_data()
    add_question_to_db(message.from_user.id, dct['full_name'], dct['question'], contact, datetime.datetime.now())
    await state.set_state(default_state)
    await state.clear()
    await message.answer(text='''
Спасибо, ваша заявка оформлена, в ближайшее время с Вами свяжутся.
    
💼 Подписывайтесь в соц. сетях, чтобы разбираться в законах без сложных терминов:
🔹 YouTube (https://www.youtube.com/@urstart)
🔹 VK (https://vk.com/club229212039)
🔹 VC (https://vc.ru/u/4590675-andrei-kuvshinov-biznes-yurist)
🔹 Дзен (https://dzen.ru/id/5de8bbf3c7e50cf95e813aaa)
    ''',
                         reply_markup=create_kb(1,
                                                new='Записаться на консультацию ✅',
                                                faq='Хочу гайд')
                         )


@router.callback_query(F.data == 'telegram', StateFilter(FSMFillForm.get_contact))
async def step_4_2(cb: types.CallbackQuery, state: FSMContext):
    add_user_to_db(
        cb.from_user.id,
        cb.from_user.username,
        cb.from_user.first_name,
        cb.from_user.last_name,
        datetime.datetime.now()
    )
    dct = await state.get_data()
    add_question_to_db(cb.from_user.id, dct['full_name'], dct['question'], 'Написать в телеграме', datetime.datetime.now())
    await state.set_state(default_state)
    await state.clear()
    await cb.message.answer(text='''
Спасибо, ваша заявка оформлена, в ближайшее время с Вами свяжутся.
    
💼 Подписывайтесь в соц. сетях, чтобы разбираться в законах без сложных терминов:
🔹 YouTube (https://www.youtube.com/@urstart)
🔹 VK (https://vk.com/club229212039)
🔹 VC (https://vc.ru/u/4590675-andrei-kuvshinov-biznes-yurist)
🔹 Дзен (https://dzen.ru/id/5de8bbf3c7e50cf95e813aaa)
    ''',
                            reply_markup=create_kb(1,
                                                   new='Записаться на консультацию ✅',
                                                   faq='Хочу гайд'))


@router.my_chat_member(ChatMemberUpdatedFilter(member_status_changed=KICKED))
async def user_blocked_bot(event: ChatMemberUpdated):
    update_user_blocked(event.from_user.id)


@router.my_chat_member(ChatMemberUpdatedFilter(member_status_changed=MEMBER))
async def user_unblocked_bot(event: ChatMemberUpdated):
    update_user_unblocked(event.from_user.id)


@router.message(F.video_note, F.from_user.id.in_(ADMIN_IDS))
async def get_note(message: types.Message):
    print(message.video_note.file_id)


@router.message(F.text == 'Excel', F.from_user.id.in_(ADMIN_IDS), StateFilter(default_state))
async def excel(message: types.Message):
    quests = get_all_questions()
    users = get_all_users()
    wb = openpyxl.Workbook()
    sh = wb['Sheet']
    for i in range(1, len(quests) + 1):
        for y in range(1, 12):
            if quests[i-1][y-1]:
                sh.cell(i, y).value = quests[i-1][y-1]
    for i in range(len(quests) + 2, len(quests) + 2 + len(users)):
        for y in range(1, 8):
            if users[i - len(quests) - 2][y-1]:
                sh.cell(i, y).value = users[i - len(quests) - 2][y-1]
    wb.save('questions.xlsx')
    await message.answer_document(FSInputFile('questions.xlsx'))


@router.message(F.text == 'Csv', F.from_user.id.in_(ADMIN_IDS), StateFilter(default_state))
async def csv(message: types.Message):
    quests = get_all_questions()
    with open('questions.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(quests)
    await message.answer_document(FSInputFile('questions.csv'))


@router.message(F.text == 'Delete_all', F.from_user.id.in_([1012882762]), StateFilter(default_state))
async def delete_all(message: Message):
    delete_all_questions()


@router.callback_query(F.data == 'faq_1', StateFilter(default_state))
async def faq(cb: CallbackQuery):
    add_user_to_db(
        cb.from_user.id,
        cb.from_user.username,
        cb.from_user.first_name,
        cb.from_user.last_name,
        datetime.datetime.now()
    )
    await cb.message.answer_document(FSInputFile('Гайд Как сохранить логистический бизнес.pdf'),
                                     reply_markup=create_kb(1,
                                                            quest_1="Записаться на консультацию ✅"))

@router.callback_query(F.data == 'faq_2', StateFilter(default_state))
async def faq(cb: CallbackQuery):
    add_user_to_db(
        cb.from_user.id,
        cb.from_user.username,
        cb.from_user.first_name,
        cb.from_user.last_name,
        datetime.datetime.now()
    )
    await cb.message.answer_document(
        FSInputFile('Гайд по Налоговым проверкам 2025.pdf'),
        reply_markup=create_kb(1,
                               quest_1="Записаться на консультацию ✅"))


@router.callback_query(F.data == 'faq_3', StateFilter(default_state))
async def faq(cb: CallbackQuery):
    add_user_to_db(
        cb.from_user.id,
        cb.from_user.username,
        cb.from_user.first_name,
        cb.from_user.last_name,
        datetime.datetime.now()
    )
    await cb.message.answer_document(
        FSInputFile('Гайд Как вести себя на допросе.pdf'),
        reply_markup=create_kb(1,
                               quest_1="Записаться на консультацию ✅"))


@router.callback_query(F.data == 'faq', StateFilter(default_state))
async def faq(cb: CallbackQuery):
    add_user_to_db(
        cb.from_user.id,
        cb.from_user.username,
        cb.from_user.first_name,
        cb.from_user.last_name,
        datetime.datetime.now()
    )
    await cb.message.answer(text="""
Мы в ЮК «СТАРТ» не просто юристы — мы ваши антикризисные союзники. Подготовили для вас 3 практичных гайда, которые помогут пережить налоговую проверку и не утонуть в кризисе.

📍 Первый гайд — антикризисное руководство для логистики: как договориться с лизингом и банками, что делать с дебиторкой, как не слить имущество и не угодить под субсидиарку.

📍 Второй — про налоговые проверки 2025: что ждать, как действовать, что говорить (а главное — чего не говорить), чтобы не оказаться в плане выездной.

📍 Третий гайд — что делать, если вызывают в налоговую или МВД, как отвечать на вопросы, чтобы не подставить себя или компанию, какие ошибки совершают даже опытные руководители и как их избежать.

🎁 Эти гайды — не вода, а конкретика и опыт. Скачивайте, читайте и действуйте. Пока другие в панике — вы уже на шаг впереди. Ссылки ниже 👇
    """,
                         reply_markup=create_kb(1,
                                                faq_1='Гайд "Как сохранить логистический бизнес"',
                                                faq_2='Гайд по Налоговым проверкам 2025',
                                                faq_3='Гайд "Как вести себя на допросе"',
                                                ))


# @router.channel_post()
# async def handle_channel_post(message: Message):
#     if message.chat.id != CHANEL_ID:
#         return
#
#     # Обработка медиагрупп
#     if message.media_group_id:
#         media_group_id = message.media_group_id
#
#         # Отменяем предыдущий таймер для этой группы
#         if media_group_id in timers:
#             timers[media_group_id].cancel()
#
#         media_groups[media_group_id].append(message)
#
#         # Запускаем новый таймер
#         timers[media_group_id] = asyncio.create_task(
#             process_media_group(media_group_id)
#         )
#     else:
#         # Обычное сообщение (не медиагруппа)
#         users = get_all_users()
#         cnt = 0
#         for user in users[1:]:
#             if not user[6]:
#                 try:
#                     await message.forward(chat_id=int(user[1]))
#                     await asyncio.sleep(0.2)
#                     cnt += 1
#                 except Exception as e:
#                     print(f"Ошибка пересылки: {e}")
#         await bot.send_message(1012882762, f'Переслано {cnt} юзерам')


#Команда для рассылки


@router.message(F.text == 'Send', StateFilter(default_state), F.from_user.id.in_(ADMIN_IDS))
async def send_to_all(message: types.Message, state: FSMContext):
    await message.answer(text='Сейчас мы подготовим сообщение для рассылки по юзерам!\n'
                              'Отправьте пжл текстовое сообщение или картинку(можно с текстом) или видео(можно с текстом) или видео-кружок')
    await state.set_state(FSMFillForm.send)


#Создание текстового сообщения


@router.message(F.text, StateFilter(FSMFillForm.send), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button(message: types.Message, state: FSMContext):
    await state.update_data(text=message.text)
    await message.answer(text='Добавим кнопку-ссылку?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.text_add_button)


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.text_add_button), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_no(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    await cb.message.answer(text='Проверьте ваше сообщение для отправки')
    await cb.message.answer(text=dct['text'])
    await cb.message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.check_text_1)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_text_1), F.from_user.id.in_(ADMIN_IDS))
async def check_text_yes_1(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            await bot.send_message(user_id, text=dct['text'])
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.text_add_button), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_yes_1(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text='Введите текст кнопки-ссылки')
    await state.set_state(FSMFillForm.text_add_button_text)


@router.message(F.text, StateFilter(FSMFillForm.text_add_button_text), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_yes_2(message: types.Message, state: FSMContext):
    await state.update_data(button_text=message.text)
    await message.answer(text='Теперь введите корректный url(ссылка на сайт, телеграмм)')
    await state.set_state(FSMFillForm.text_add_button_url)


@router.message(F.text, StateFilter(FSMFillForm.text_add_button_url), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_yes_3(message: types.Message, state: FSMContext):
    await state.update_data(button_url=message.text)
    dct = await state.get_data()
    try:
        await message.answer(text='Проверьте ваше сообщение для отправки')
        await message.answer(text=dct['text'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        await message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
        await state.set_state(FSMFillForm.check_text_2)
    except Exception:
        await message.answer(text='Скорее всего вы ввели не корректный url. Направьте корректный url')
        await state.set_state(FSMFillForm.text_add_button_url)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_text_2), F.from_user.id.in_(ADMIN_IDS))
async def check_text_yes_2(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            await bot.send_message(user_id, text=dct['text'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.check_text_1, FSMFillForm.check_text_2), F.from_user.id.in_(ADMIN_IDS))
async def check_message_no(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text=f'Сообщение не отправлено')
    await state.set_state(default_state)
    await state.clear()


#Создание фото-сообщения


@router.message(F.photo, StateFilter(FSMFillForm.send), F.from_user.id.in_(ADMIN_IDS))
async def photo_add_button(message: types.Message, state: FSMContext):
    await state.update_data(photo_id=message.photo[-1].file_id)
    try:
        await state.update_data(caption=message.caption)
    except Exception:
        pass
    await message.answer(text='Добавим кнопку-ссылку?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.photo_add_button)


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.photo_add_button), F.from_user.id.in_(ADMIN_IDS))
async def text_add_button_no(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    await cb.message.answer(text='Проверьте ваше сообщение для отправки')
    if dct.get('caption'):
        await cb.message.answer_photo(photo=dct['photo_id'], caption=dct['caption'])
    else:
        await cb.message.answer_photo(photo=dct['photo_id'])
    await cb.message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.check_photo_1)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_photo_1), F.from_user.id.in_(ADMIN_IDS))
async def check_photo_yes_1(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            if dct.get('caption'):
                await bot.send_photo(user_id, photo=dct['photo_id'], caption=dct['caption'])
            else:
                await bot.send_photo(user_id, photo=dct['photo_id'])
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.photo_add_button), F.from_user.id.in_(ADMIN_IDS))
async def photo_add_button_yes_1(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text='Введите текст кнопки-ссылки')
    await state.set_state(FSMFillForm.photo_add_button_text)


@router.message(F.text, StateFilter(FSMFillForm.photo_add_button_text), F.from_user.id.in_(ADMIN_IDS))
async def photo_add_button_yes_2(message: types.Message, state: FSMContext):
    await state.update_data(button_text=message.text)
    await message.answer(text='Теперь введите корректный url(ссылка на сайт, телеграмм)')
    await state.set_state(FSMFillForm.photo_add_button_url)


@router.message(F.text, StateFilter(FSMFillForm.photo_add_button_url), F.from_user.id.in_(ADMIN_IDS))
async def photo_add_button_yes_3(message: types.Message, state: FSMContext):
    await state.update_data(button_url=message.text)
    dct = await state.get_data()
    try:
        await message.answer(text='Проверьте ваше сообщение для отправки')
        if dct.get('caption'):
            await message.answer_photo(photo=dct['photo_id'], caption=dct['caption'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        else:
            await message.answer_photo(photo=dct['photo_id'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        await message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
        await state.set_state(FSMFillForm.check_photo_2)
    except Exception as e:
        print(e)
        await message.answer(text='Скорее всего вы ввели не корректный url. Направьте корректный url')
        await state.set_state(FSMFillForm.photo_add_button_url)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_photo_2), F.from_user.id.in_(ADMIN_IDS))
async def check_photo_yes_2(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            if dct.get('caption'):
                    await bot.send_photo(user_id, photo=dct['photo_id'], caption=dct['caption'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            else:
                await bot.send_photo(user_id, photo=dct['photo_id'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.check_text_1, FSMFillForm.check_text_2,
            FSMFillForm.check_photo_1, FSMFillForm.check_photo_2), F.from_user.id.in_(ADMIN_IDS))
async def check_message_no(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text=f'Сообщение не отправлено')
    await state.set_state(default_state)
    await state.clear()


#Создание видео-сообщения


@router.message(F.video, StateFilter(FSMFillForm.send), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button(message: types.Message, state: FSMContext):
    await state.update_data(video_id=message.video.file_id)
    try:
        await state.update_data(caption=message.caption)
    except Exception:
        pass
    await message.answer(text='Добавим кнопку-ссылку?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.video_add_button)


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.video_add_button), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button_no(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    await cb.message.answer(text='Проверьте ваше сообщение для отправки')
    if dct.get('caption'):
        await cb.message.answer_video(video=dct['video_id'], caption=dct['caption'])
    else:
        await cb.message.answer_video(video=dct['video_id'])
    await cb.message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.check_video_1)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_video_1), F.from_user.id.in_(ADMIN_IDS))
async def check_video_yes_1(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            if dct.get('caption'):
                await bot.send_video(user_id, video=dct['video_id'], caption=dct['caption'])
            else:
                await bot.send_video(user_id, video=dct['video_id'])
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.video_add_button), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button_yes_1(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text='Введите текст кнопки-ссылки')
    await state.set_state(FSMFillForm.video_add_button_text)


@router.message(F.text, StateFilter(FSMFillForm.video_add_button_text), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button_yes_2(message: types.Message, state: FSMContext):
    await state.update_data(button_text=message.text)
    await message.answer(text='Теперь введите корректный url(ссылка на сайт, телеграмм)')
    await state.set_state(FSMFillForm.video_add_button_url)


@router.message(F.text, StateFilter(FSMFillForm.video_add_button_url), F.from_user.id.in_(ADMIN_IDS))
async def video_add_button_yes_3(message: types.Message, state: FSMContext):
    await state.update_data(button_url=message.text)
    dct = await state.get_data()
    try:
        await message.answer(text='Проверьте ваше сообщение для отправки')
        if dct.get('caption'):
            await message.answer_video(video=dct['video_id'], caption=dct['caption'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        else:
            await message.answer_video(video=dct['video_id'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
        await message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
        await state.set_state(FSMFillForm.check_video_2)
    except Exception as e:
        print(e)
        await message.answer(text='Скорее всего вы ввели не корректный url. Направьте корректный url')
        await state.set_state(FSMFillForm.video_add_button_url)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_video_2), F.from_user.id.in_(ADMIN_IDS))
async def check_video_yes_2(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            if dct.get('caption'):
                await bot.send_video(user_id, video=dct['video_id'], caption=dct['caption'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            else:
                await bot.send_video(user_id, photo=dct['video_id'], reply_markup=kb_button(dct['button_text'], dct['button_url']))
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


#Создание видео-кружка


@router.message(F.video_note, StateFilter(FSMFillForm.send), F.from_user.id.in_(ADMIN_IDS))
async def video_note_check(message: types.Message, state: FSMContext):
    await state.update_data(video_note_id=message.video_note.file_id)
    await message.answer(text='Проверьте вашу запись в кружке для отправки')
    await message.answer(text='Отправляем?', reply_markup=create_kb(2, yes='Да', no='Нет'))
    await state.set_state(FSMFillForm.check_video_note_1)


@router.callback_query(F.data == 'yes', StateFilter(FSMFillForm.check_video_note_1), F.from_user.id.in_(ADMIN_IDS))
async def check_video_note_yes_1(cb: types.CallbackQuery, state: FSMContext):
    dct = await state.get_data()
    users = get_all_users_unblock()
    count = 0
    for user_id in users:
        try:
            await bot.send_video_note(user_id, video_note=dct['video_note_id'])
            count += 1
        except Exception as e:
            await bot.send_message(1012882762, str(e))
            await bot.send_message(1012882762, str(user_id))
    await cb.message.answer(text=f'Сообщение отправлено {count} юзерам')
    await state.set_state(default_state)
    await state.clear()


# Выход из рассылки без отправки


@router.callback_query(F.data == 'no', StateFilter(FSMFillForm.check_text_1, FSMFillForm.check_text_2,
                       FSMFillForm.check_photo_1, FSMFillForm.check_photo_2, FSMFillForm.check_video_1,
                       FSMFillForm.check_video_2, FSMFillForm.check_video_note_1), F.from_user.id.in_(ADMIN_IDS))
async def check_message_no(cb: types.CallbackQuery, state: FSMContext):
    await cb.message.answer(text=f'Сообщение не отправлено')
    await state.set_state(default_state)
    await state.clear()





#Рассылка текста одному юзеру
@router.message(F.text == 'Sendid', StateFilter(default_state), F.from_user.id.in_(ADMIN_IDS))
async def send_to_one_1(message: types.Message, state: FSMContext):
    await message.answer(text='Введите id юзера')
    await state.set_state(FSMFillForm.send_id)


@router.message(F.text, StateFilter(FSMFillForm.send_id), F.from_user.id.in_(ADMIN_IDS))
async def send_to_one_2(message: types.Message, state: FSMContext):
    try:
        await state.update_data(user_id=int(message.text))
        await message.answer(text='Введите сообщение для отправки юзеру по id')
        await state.set_state(FSMFillForm.send_to_one)
    except Exception:
        await message.answer(text='Что-то пошло не так, проверьте корректность id. Попробуйте снова')
        await state.set_state(default_state)


@router.message(F.text, StateFilter(FSMFillForm.send_to_one), F.from_user.id.in_(ADMIN_IDS))
async def send_to_one_3(message: types.Message, state: FSMContext):
    try:
        dct = await state.get_data()
        await bot.send_message(chat_id=dct['user_id'], text=message.text)
        await message.answer(text=f"Сообщение юзеру с id {dct['user_id']} отправлено")
    except Exception:
        await message.answer(text='Что-то пошло не так, проверьте корректность id или блокировку бота юзером. Попробуйте снова')
    await state.set_state(default_state)
    await state.clear()


#Пересылка сообщений от клиента админам
@router.message(F.text, ~F.from_user.id.in_(ADMIN_IDS))
async def forward_message(message: types.Message):
    for admin_id in ADMIN_IDS:
        try:
            await bot.forward_message(chat_id=admin_id, from_chat_id=message.chat.id, message_id=message.message_id)
            await bot.send_message(chat_id=admin_id, text=f'{message.from_user.username}(ID{message.from_user.id})')
        except:
            pass


#Ответ на сообщения юзера админом через reply

@router.message(F.text, F.from_user.id.in_(ADMIN_IDS))
async def answer_admin_text(message: Message):
    if message.reply_to_message:
        try:
            if '(ID' in message.reply_to_message.text:
                user_id = int(message.reply_to_message.text[:-1].split('(ID')[-1])
                await bot.send_message(chat_id=user_id, text=message.text)
                await message.answer(text=f'Сообщение юзеру с id {user_id} отправлено')
        except Exception:
            await message.answer(text='Что-то пошло не так')
